
import time
import os
import glob
import signal
from openpyxl import Workbook, load_workbook
from synth import Synthesizer
import bu_astar 
import td_astar 
import argparse
import sys

def timeout_handler(signum, frame):
    """
    Signal handler that raises a TimeoutError. 
    This is triggered by calling `signal.alarm(seconds)`.
    """
    raise TimeoutError

class Runner:
    """
    The Runner class orchestrates the entire process of synthesizing solutions 
    for one or more benchmarks. It sets up a timeout, runs a Synthesizer instance, 
    and records the results in an Excel file.

    Attributes:
        directory_path (str): The path where benchmark files and logs are located.
        clang_path (str): The path to the Clang executable for code analysis.
        nn_solution (int): The number of lines to consider if using neural network solutions (for pre-check).
        grammar_style (str): The grammar style used ('wcfg', 'wcfg_equal_p', 'full_grammar', etc.).
        benchmark (str): The particular benchmark name (e.g., 'dot', 'matmul').
        timeout (int): How many seconds to allow before timing out the synthesis.
        enumerator (str): Which enumerator method to use ('top_down' or 'bottom_up').
        pre_check (bool): Whether to check on LLM solutions.
        debug (bool): If True, does not update Excel logs and provides additional printouts.
        excel_file (str): Path to the Excel file used for logging synthesis results.
        analyzer_synthesizer (Synthesizer): An instance of the Synthesizer class handling compilation and checks.
    """

    def __init__(self, directory_path, clang_path, nn_solution, grammar_style, benchmark, timeout, enumerator, pre_check, debug=False):
        """
        Initializes the Runner by storing parameters, setting up the analyzer 
        and ensuring the Excel file for logging exists.

        Args:
            directory_path (str): Path to the benchmark/data directory.
            clang_path (str): Path to the clang executable for code analysis.
            nn_solution (int): Number of lines for neural net-based solutions (if used).
            grammar_style (str): Grammar approach.
            benchmark (str): The name of the benchmark to run.
            timeout (int): Allowed runtime in seconds before a TimeoutError is raised.
            enumerator (str): Enumerator style ('top_down', 'bottom_up').
            pre_check (bool): Flag for performing a check on LLM solutions.
            debug (bool): If True, skip updating Excel and enable verbose prints.
        """
        self.directory_path = directory_path
        self.clang_path = clang_path
        self.nn_solution = nn_solution
        self.grammar_style = grammar_style
        self.timeout = timeout
        self.enumerator = enumerator
        self.pre_check = pre_check
        self.benchmark = benchmark
        self.debug = debug
        print(f'Running with enumerator: {self.enumerator}')
        if self.grammar_style == 'wcfg':
            print('Using WCFG')
        if self.grammar_style == 'wcfg_equal_p':
            print('Using wcfg with even weights')
        if self.grammar_style == 'full_grammar':
            print('Using full grammar')
        if self.grammar_style == 'original':
            print('Using original grammar')
        if self.pre_check:
            check = 'pre_check'
        else:
            check = 'no_pre_check'
        self.benchmark_files = self._get_benchmark_files(benchmark)
        self.analyzer_synthesizer = Synthesizer(self.directory_path, self.clang_path)
        if not os.path.exists(f"{self.directory_path}/lifting_logs"):
            os.makedirs(f"{self.directory_path}/lifting_logs")
        self.excel_file = f'{self.directory_path}/lifting_logs/{enumerator}_{grammar_style}_{check}.xlsx'
        self._init_excel_file()

    def _get_benchmark_files(self, benchmark):
        """
        Retrieve benchmark file paths. If 'benchmark' is None, 
        it gathers all .c files within 'benchmarks/'. Otherwise, returns 
        the single specified benchmark.

        Args:
            benchmark (str): The benchmark name or None.

        Returns:
            list of str: Sorted list of file paths to benchmarks.
        """
        if benchmark is None:
            # Get all benchmark files and sort them alphabetically
            files = glob.glob(f'{self.directory_path}/benchmarks/*.c')
            return sorted(files)
        else:
            # Return the single benchmark file in a list (sorting is not necessary for a single item)
            return [f'{self.directory_path}/benchmarks/{benchmark}.c']


    def _init_excel_file(self):
        """
        Initialize the Excel file for logging results if it does not already exist.
        Creates a header row: ['Benchmark', 'Grammar Style', 'Total Synthesis Time', 
                               'Candidates Tried', 'Solution', 'Synth Method'].
        """
        if not os.path.exists(self.excel_file):
            wb = Workbook()
            ws = wb.active
            ws.append(['Benchmark', 'Grammar Style', 'Total Synthesis Time', 
                        'Candidates Tried', 'Solution', 'Synth Method'])
            wb.save(self.excel_file)

    def run(self):
        """
        Main entry point to run synthesis on a chosen benchmark. Sets up a timeout, 
        logs the start time, and invokes the synthesizer. On success or timeout, 
        logs the result into the Excel spreadsheet (unless in debug mode).
        """
        # benchmark = [os.path.basename(f).replace('.c', '') for f in self.benchmark_files][0]

        signal.signal(signal.SIGALRM, timeout_handler)
        signal.alarm(self.timeout)  # Set timeout
        total_syn_time_start = time.time()
        print(f'Processing: {self.benchmark}')
        candidates_tried = 0
        solution = 'no solution'
        synth_method = self.enumerator
        # Setting up the timeout
        try:
            solution, candidates_tried, synth_method = self.analyzer_synthesizer.synthesize(self.benchmark, self.nn_solution, self.grammar_style, self.enumerator, self.pre_check)
            total_syn_time = time.time() - total_syn_time_start
            if self.debug == False:
                self._update_excel(self.benchmark, total_syn_time, candidates_tried, solution, synth_method)
        except TimeoutError:
            candidates_tried = td_astar.candidates_tried_td if self.enumerator == 'top_down' else bu_astar.candidates_tried_bu
            print(f'Benchmark {self.benchmark} exceeded the runtime limit of {self.timeout} seconds.')
            print(f'candidates tried: {candidates_tried}')
            total_syn_time = time.time() - total_syn_time_start
            if self.debug == False:
                self._update_excel(self.benchmark, 'Timeout', candidates_tried, 'no solution','Failed')
        finally:
            signal.alarm(0)
        print(f'wcfg status: {self.grammar_style}')
        print(f'Candidates tried: {candidates_tried}')
        print(f'Total synthesis time: {total_syn_time}')
        print(f'Solution: {solution}')
        print(f'Synth Method: {synth_method}\n')

    def _update_excel(self, benchmark, timeout, total_syn_time, candidates_tried, solution=None, synth_method=None):
        """
        Log the results of a benchmark run to the Excel file. Appends a row containing:
          [benchmark, grammar_style, timeout, total_syn_time, candidates_tried, solution, synth_method]

        Args:
            benchmark (str): Name of the benchmark processed.
            timeout (str | float): Either the numeric time or 'Timeout' string.
            total_syn_time (float | str): Numeric time or 'Timeout' string.
            candidates_tried (int): How many candidates were enumerated.
            solution (str | None): The solution expression if found, else 'no solution'.
            synth_method (str | None): The enumerator or method used.
        """
        if not os.path.exists(self.excel_file):
            self._init_excel_file()
        wb = load_workbook(self.excel_file)
        ws = wb.active
        ws.append([benchmark, self.grammar_style, timeout, total_syn_time, candidates_tried, solution, synth_method])
        wb.save(self.excel_file)



def str_to_bool(value):
    """
    Convert a string to a boolean. Accepts variants of 
    'yes', 'true', 't', '1' -> True, and
    'no', 'false', 'f', '0' -> False.

    Args:
        value (str|bool): The string or bool to convert.

    Returns:
        bool: The corresponding boolean value.

    Raises:
        argparse.ArgumentTypeError: If the string is not a recognized boolean form.
    """
    if isinstance(value, bool):
        return value
    if value.lower() in ('yes', 'true', 't', '1'):
        return True
    elif value.lower() in ('no', 'false', 'f', '0'):
        return False
    else:
        raise argparse.ArgumentTypeError('Boolean value expected.')
    
if __name__ == '__main__':
    directory_path = 'data'
    clang_path = '../llvm/bin/clang'
    nn_solution = 99
    parser = argparse.ArgumentParser(description='Run the Lifting tool on all benchmarks')
    parser.add_argument('--grammar_style', type=str, default='wcfg', help='The grammar style to use')
    parser.add_argument('--benchmark', type=str, default=None, help='The benchmark to run on.')
    parser.add_argument('--timeout', type=int, default=10*60, help='The timeout for each benchmark')
    parser.add_argument('--enumerator', type=str, default='top_down', help='The enumerator to use')
    parser.add_argument('--check_llm', type=str_to_bool, default=False, help='Whether to check LLM solutions (true/false)')
    
    args = parser.parse_args()
    if len(sys.argv) !=1:
        try:
            runner = Runner(directory_path, clang_path, nn_solution, args.grammar_style, args.benchmark, args.timeout, args.enumerator, args.check_llm)
            runner.run()
        except:
            print("Error: Invalid command-line arguments provided.")
            sys.exit(1)
    else:
        print('Debugging mode')
        grammar_style = 'wcfg'
        benchmark = "dot" 
        timeout = 60 * 60
        enumerator = 'top_down'
        # enumerator = 'bottom_up'
        check_llm = False
        benchmark_runner = Runner(directory_path, clang_path, nn_solution, grammar_style, benchmark, timeout, enumerator, check_llm, debug=True)
        benchmark_runner.run()
